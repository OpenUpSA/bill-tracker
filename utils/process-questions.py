#!/usr/bin/env python3
"""
Process the PMG Questions Register Excel + PMG questions export CSV
into a consolidated questions CSV with PMG question IDs, member IDs, and
departments.

Usage:
    python3 utils/process-questions.py <register.xlsx> <pmg-export.csv> \
        [--members <current-members.csv>] [--out <output.csv>] [--report <report.txt>]

Example (2025):
    python3 utils/process-questions.py \
        "src/data/questions/2025/REGISTER 2025_cleaned.xlsx" \
        "src/data/questions/2025/questions-2025.csv" \
        --members "src/data/questions/current-members.csv" \
        --out "src/data/questions/2025/questions-2025-all.csv" \
        --report "src/data/questions/2025/questions-2025-match-report.txt"

Example (2026):
    python3 utils/process-questions.py \
        "src/data/questions/2026/REGISTER 2026 RECONSTRUCTION_cleaned.xlsx" \
        "src/data/questions/2026/questions-2026.csv" \
        --members "src/data/questions/2026/current-members.csv" \
        --out "src/data/questions/2026/questions-2026-all.csv" \
        --report "src/data/questions/2026/questions-2026-match-report.txt"

Reads the WRITTENS, ORALS, PRESIDENT, and DEPUTY PRESIDENT sheets from the
Excel workbook, combines them into one CSV, and adds:

  * `pmg_id`        – PMG question ID (register ACC No, trailing 'E' stripped,
                      matched to PMG export Code).
  * `member_id`     – PMG member ID of the MP who asked the question.
                      Found from the PMG export when the Code matches, or by
                      fuzzy name matching against current-members.csv when the
                      Code does not match.
  * `department`    – the Executive (portfolio) the question was directed to,
                      taken from the register's `Executive` column.
  * `member_match`  – how `member_id` was determined:
                      'pmg_code'     – matched via PMG export Code.
                      'pmg_code_name'– Code matched and name confirmed/tie-broken.
                      'member_name'  – fuzzy name match against current-members.
                      'ambiguous'    – multiple equally-good member candidates.
                      'no_match'     – could not confidently match a member.
                      'empty_member' – register has no member name to match.

A `type` column records the source sheet: 'written', 'oral',
'president', or 'deputy_president'.

A match report is written listing every row that was ambiguous or unmatched,
so a human can review and fill in the member_id manually.
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl is required. Install with: pip3 install openpyxl")


# ── Helpers ──────────────────────────────────────────────────────────────────

TITLES = {"dr", "mr", "ms", "mrs", "miss", "prof", "rev", "hon", "adv", "inkosi"}
PARTICLES = {
    "du", "le", "te", "de", "da", "di", "el", "al", "la", "van", "der",
    "den", "ten", "von", "del", "della", "ben", "ibn", "st", "ste", "of",
}


def fmt(v):
    """Format an Excel cell value for CSV output."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def strip_e(s):
    """Strip trailing 'E' from ACC No to match PMG Code. e.g. 'NW1E' -> 'NW1'"""
    return s[:-1] if s and s.endswith("E") else s


def parse_name(s):
    """Split a name into (title, initials, name_tokens).

    Handles both register format ('Ms N A Gcaleka-Mazibuko') and member format
    ('Gcaleka-Mazibuko, Ms N A'). The title (Dr, Mr, Ms, Mrs, ...) is extracted
    first. Hyphenated surnames are split into their parts. Multi-letter
    all-caps tokens of <=4 chars are treated as initials (e.g. 'SG' -> s, g).
    Particles (du, le, van, ...) are kept as name tokens.
    """
    s = (s or "").strip()
    s = re.sub(r"^[^A-Za-z]+", "", s)  # strip leading non-alpha (e.g. '1600 ')
    s = s.replace(",", " ").replace(".", " ")
    raw = re.findall(r"[A-Za-z]+(?:-[A-Za-z]+)*", s)
    # Extract the first title token (case-insensitive)
    title = ""
    toks = []
    for t in raw:
        tl = t.lower()
        if not title and tl in TITLES:
            title = tl
        else:
            toks.append(t)
    initials = set()
    name_toks = set()
    for t in toks:
        for p in t.split("-"):
            pl = p.lower()
            if len(p) == 1:
                initials.add(pl)
            elif p.isupper() and len(p) <= 4 and pl not in PARTICLES:
                for c in pl:
                    initials.add(c)
            elif pl in PARTICLES:
                name_toks.add(pl)
            else:
                name_toks.add(pl)
    return title, frozenset(initials), frozenset(name_toks)


# Title groups: titles that share the same implied gender / are interchangeable
# for matching purposes. 'dr' and 'prof' are gender-neutral and compatible
# with anything. 'rev', 'hon', 'adv', 'inkosi' are also gender-neutral.
TITLE_GROUPS = [
    {"ms", "mrs", "miss"},
    {"mr"},
]
GENDER_NEUTRAL_TITLES = {"dr", "prof", "rev", "hon", "adv", "inkosi"}


def title_compatible(t1, t2):
    """True if two titles are compatible (same group or one is gender-neutral)."""
    if not t1 or not t2:
        return True  # missing title = no constraint
    if t1 == t2:
        return True
    if t1 in GENDER_NEUTRAL_TITLES or t2 in GENDER_NEUTRAL_TITLES:
        return True
    for grp in TITLE_GROUPS:
        if t1 in grp and t2 in grp:
            return True
    return False


def init_compatible(a, b):
    """True if initial sets are compatible (one is a subset of the other)."""
    if not a or not b:
        return True
    return a <= b or b <= a


def name_compatible(a, b):
    """True if surname token sets are compatible.

    Exact match, subset (one has extra middle names), or share at least one
    token with at most one differing token (handles hyphenation differences).
    """
    if not a or not b:
        return False
    if a == b or a <= b or b <= a:
        return True
    return len(a & b) >= 1 and len(a ^ b) <= 1


# ── Sheet config ─────────────────────────────────────────────────────────────
# Each sheet has 2 title/legend rows, then the header on row 3 (0-indexed 2).
# Sheets are listed in output order.

SHEETS = [
    {"name": "WRITTENS",          "type": "written"},
    {"name": "ORALS",             "type": "oral"},
    {"name": "PRESIDENT",         "type": "president"},
    {"name": "DEPUTY PRESIDENT",  "type": "deputy_president"},
]


def read_sheet(ws):
    """Read a worksheet, returning (header_list, data_rows).

    Assumes rows 1-2 are title/legend, row 3 is the column header, and
    data starts at row 4. Trailing empty columns are trimmed.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        return [], []

    header = [fmt(c) for c in all_rows[2]]
    while header and header[-1] == "":
        header.pop()
    ncols = len(header)

    data = []
    for r in all_rows[3:]:
        if all(c is None for c in r):
            continue
        data.append([fmt(r[i]) if i < len(r) else "" for i in range(ncols)])

    return header, data


def load_pmg(csv_path):
    """Load PMG export CSV and index by stripped Code.

    Returns dict: code -> list of PMG row dicts.
    """
    csv.field_size_limit(sys.maxsize)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    by_key = defaultdict(list)
    for r in rows:
        code = r.get("Code", "")
        if code:
            by_key[code].append(r)
    return by_key


def load_members(csv_path):
    """Load current-members CSV and build a search index by surname token.

    Returns list of ((title, initials, name_tokens), member_row) tuples.
    """
    if not csv_path:
        return []
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    return [(parse_name(r["Name"]), r) for r in rows]


def build_member_index(members):
    """Index members by each of their surname tokens for fast lookup."""
    idx = defaultdict(list)
    for parsed, m in members:
        _, _, name_toks = parsed
        for t in name_toks:
            idx[t].append((parsed, m))
    return idx


def match_pmg_question(acc_no, member, pmg_by_key):
    """Match a register row to a PMG question row by Code (trailing E stripped).

    Returns (pmg_id, member_id, status) where status is 'pmg_code' (single
    match) or 'pmg_code_name' (tie-broken among duplicate codes by name), or
    ("", "", None) if no Code match.
    """
    key = strip_e(acc_no)
    if not key or key not in pmg_by_key:
        return "", "", None
    candidates = pmg_by_key[key]
    if len(candidates) == 1:
        c = candidates[0]
        return c.get("ID", ""), c.get("Asked By Member ID", ""), "pmg_code"
    # Multiple PMG rows with same code — tie-break by member name compatibility
    rt, ri, rn = parse_name(member)
    best = None
    best_score = -1
    for c in candidates:
        ct, ci, cn = parse_name(c.get("Asked By Name", ""))
        score = (2 if name_compatible(rn, cn) else 0) + (
            2 if init_compatible(ri, ci) else 0
        ) + (1 if title_compatible(rt, ct) else 0)
        if score > best_score:
            best_score = score
            best = c
    return (
        best.get("ID", "") if best else "",
        best.get("Asked By Member ID", "") if best else "",
        "pmg_code_name",
    )


def match_member_by_name(member, member_index):
    """Fuzzy-match a register member name against current-members.

    Returns (member_id, status, candidates) where status is 'member_name',
    'ambiguous', 'no_match', or 'empty_member'. `candidates` is the list of
    candidate member rows (for reporting).

    Title (Dr, Mr, Ms, ...) is used both as a filter (incompatible titles are
    excluded) and as a scoring boost, so 'Ms P Marais' prefers 'Marais, Ms P'
    over 'Marais, Mr GP'.
    """
    rt, ri, rn = parse_name(member)
    if not rn:
        return "", "empty_member", []

    # Gather members sharing any surname token, filtered by compatibility
    seen = set()
    scored = []
    for t in rn:
        for (parsed, m) in member_index.get(t, []):
            if id(m) in seen:
                continue
            seen.add(id(m))
            mt, mi, mn = parsed
            if name_compatible(rn, mn) and init_compatible(ri, mi):
                # Base score from name/initial overlap, plus title agreement
                score = len(rn & mn) + len(ri & mi)
                if title_compatible(rt, mt):
                    score += 2  # title-compatible candidates rank higher
                else:
                    score -= 2  # title mismatch penalized but not excluded
                scored.append((score, m))

    if not scored:
        return "", "no_match", []
    scored.sort(key=lambda x: -x[0])

    # If the top two have the same score, we cannot confidently pick one
    if len(scored) > 1 and scored[0][0] == scored[1][0]:
        return "", "ambiguous", [m for _, m in scored[:5]]

    return scored[0][1].get("ID", ""), "member_name", [scored[0][1]]


def prompt_choice(member_name, cands):
    """Interactively ask the user to choose the correct member from candidates.

    Returns the chosen member row, or None if the user skips.
    """
    print(f"\n  Ambiguous match for: '{member_name}'")
    print(f"  Multiple members could match. Please choose:")
    for i, c in enumerate(cands):
        print(f"    [{i + 1}]  ID={c.get('ID'):>5s}  {c.get('Name')}")
    print(f"    [s]  skip (leave member_id blank, report as ambiguous)")
    while True:
        try:
            choice = input("  > ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\n  (skipped)")
            return None
        if choice == "s":
            return None
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(cands):
                return cands[idx]
        print("  Please enter a number or 's' to skip.")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert PMG Questions Register Excel + PMG export CSV "
                    "into a consolidated questions CSV with PMG IDs, member "
                    "IDs, and departments."
    )
    parser.add_argument(
        "register",
        help="Path to the Register Excel file (.xlsx)",
    )
    parser.add_argument(
        "pmg_export",
        help="Path to the PMG questions export CSV",
    )
    parser.add_argument(
        "--members", "-m",
        default=None,
        help="Path to current-members CSV for fuzzy name matching "
             "(default: src/data/questions/2026/current-members.csv if present)",
    )
    parser.add_argument(
        "--out", "-o",
        default="questions-all.csv",
        help="Output CSV path (default: questions-all.csv)",
    )
    parser.add_argument(
        "--report", "-r",
        default=None,
        help="Write a human-readable match report (ambiguous + unmatched) "
             "to this path",
    )
    parser.add_argument(
        "--no-interactive", "--no-prompt",
        action="store_true",
        default=False,
        help="Don't interactively prompt for ambiguous matches; just report "
             "them (useful for batch/cron runs)",
    )
    args = parser.parse_args()

    # Default members path if not specified
    members_path = args.members
    if members_path is None:
        default_m = "src/data/questions/2026/current-members.csv"
        import os
        if os.path.exists(default_m):
            members_path = default_m

    # Load PMG export
    print(f"Loading PMG export: {args.pmg_export}")
    pmg_by_key = load_pmg(args.pmg_export)
    print(f"  {sum(len(v) for v in pmg_by_key.values())} PMG rows, "
          f"{len(pmg_by_key)} unique codes")

    # Load members for fuzzy matching
    member_index = {}
    if members_path:
        print(f"Loading members: {members_path}")
        members = load_members(members_path)
        member_index = build_member_index(members)
        print(f"  {len(members)} members indexed")
    else:
        print("No members CSV provided – Code-unmatched rows will have no member_id")

    # Load workbook
    print(f"Loading register: {args.register}")
    wb = openpyxl.load_workbook(args.register, data_only=True, read_only=True)

    # Collect rows from all sheets.
    # Final columns: type, <all sheet columns>, pmg_id, member_id, department,
    # member_match
    sheet_data = {}       # sheet_type -> (header, rows)
    all_headers_seen = []  # ordered list of all column names across sheets

    for s in SHEETS:
        if s["name"] not in wb.sheetnames:
            print(f"  ⚠️  Sheet '{s['name']}' not found, skipping")
            continue
        ws = wb[s["name"]]
        header, data = read_sheet(ws)
        sheet_data[s["type"]] = (header, data)
        print(f"  Sheet '{s['name']}': {len(data)} rows, {len(header)} cols")
        for h in header:
            if h not in all_headers_seen:
                all_headers_seen.append(h)

    # Build unified output
    out_header = (
        ["type"]
        + all_headers_seen
        + ["pmg_id", "member_id", "department", "member_match"]
    )

    out_rows = []
    report_lines = []  # lines for the human-readable match report
    stats = defaultdict(lambda: {"rows": 0, "pmg_id": 0, "member_id": 0})
    match_status_counts = defaultdict(int)

    # Cache for interactive choices: member_name -> chosen member_id
    # so we only ask once per unique ambiguous name.
    manual_choices = {}
    interactive = not args.no_interactive
    if interactive:
        print("\nInteractive mode: ambiguous matches will prompt for a choice.")
        print("(Use --no-interactive to skip prompts and just report them.)\n")

    for s in SHEETS:
        sd = sheet_data.get(s["type"])
        if not sd:
            continue
        header, data = sd
        acc_idx = header.index("ACC No") if "ACC No" in header else None
        member_idx = header.index("Member asking") if "Member asking" in header else None
        exec_idx = header.index("Executive") if "Executive" in header else None
        qno_idx = header.index("Ques No.") if "Ques No." in header else None

        for row in data:
            acc = row[acc_idx] if acc_idx is not None else ""
            member = row[member_idx] if member_idx is not None else ""
            execv = row[exec_idx] if exec_idx is not None else ""
            qno = row[qno_idx] if qno_idx is not None else ""

            pmg_id = ""
            member_id = ""
            status = "no_match"

            # 1. Try Code match against PMG export
            pmg_id, pmg_mid, pmg_status = match_pmg_question(
                acc, member, pmg_by_key
            )
            if pmg_status:
                member_id = pmg_mid
                status = pmg_status

            # 2. If no Code match (or Code match but no member_id), try name
            if not member_id and member_index:
                nm_id, nm_status, cands = match_member_by_name(
                    member, member_index
                )
                if nm_status == "member_name":
                    member_id = nm_id
                    status = "member_name"
                elif nm_status == "ambiguous":
                    # Try interactive disambiguation (cached per unique name)
                    if interactive and member not in manual_choices:
                        chosen = prompt_choice(member, cands)
                        if chosen:
                            manual_choices[member] = chosen.get("ID", "")
                            print(f"  -> recorded: '{member}' = "
                                  f"{chosen.get('ID')} ({chosen.get('Name')})")
                        else:
                            manual_choices[member] = ""  # skipped
                    chosen_id = manual_choices.get(member, "")
                    if chosen_id:
                        member_id = chosen_id
                        status = "manual"
                    else:
                        status = "ambiguous"
                        cand_names = ", ".join(
                            f"{c.get('ID')}: {c.get('Name')}" for c in cands
                        )
                        report_lines.append(
                            f"AMBIGUOUS  [{s['type']}] Q{qno} ACC={acc}  "
                            f"'{member}' -> candidates: {cand_names}"
                        )
                elif nm_status in ("no_match", "empty_member"):
                    status = nm_status
                # Report no_match rows for human review
                if nm_status == "no_match":
                    report_lines.append(
                        f"NO_MATCH   [{s['type']}] Q{qno} ACC={acc}  "
                        f"'{member}' -> no member found"
                    )
            elif status in ("pmg_code", "pmg_code_name") and not member_id:
                # Code matched but PMG has no member_id for this row — note it
                report_lines.append(
                    f"NO_MEMBER_ID [{s['type']}] Q{qno} ACC={acc}  "
                    f"'{member}' -> PMG Code matched but no member_id"
                )

            match_status_counts[status] += 1

            row_map = dict(zip(header, row))
            out_row = {
                "type": s["type"],
                "pmg_id": pmg_id,
                "member_id": member_id,
                "department": execv,
                "member_match": status,
            }
            for h in all_headers_seen:
                out_row[h] = row_map.get(h, "")
            out_rows.append(out_row)

        stats[s["type"]]["rows"] = len(data)
        stats[s["type"]]["pmg_id"] = sum(
            1 for r in out_rows[-len(data):] if r["pmg_id"]
        )
        stats[s["type"]]["member_id"] = sum(
            1 for r in out_rows[-len(data):] if r["member_id"]
        )

    # Write CSV
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=out_header, extrasaction="ignore")
        w.writeheader()
        w.writerows(out_rows)

    # Write match report
    if args.report:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(
                f"PMG Questions Match Report\n"
                f"==========================\n\n"
                f"Total rows: {len(out_rows)}\n"
            )
            for status_name in [
                "pmg_code", "pmg_code_name", "member_name", "manual",
                "ambiguous", "no_match", "empty_member",
            ]:
                f.write(
                    f"  {status_name:16s}: "
                    f"{match_status_counts.get(status_name, 0)}\n"
                )
            f.write(f"\n--- Rows needing review ({len(report_lines)}) ---\n\n")
            for line in report_lines:
                f.write(line + "\n")
        print(f"Wrote match report: {args.report} "
              f"({len(report_lines)} rows need review)")

    # Console report
    print(f"\nWrote {len(out_rows)} rows to {args.out}")
    print(f"  columns: {out_header}\n")
    for s in SHEETS:
        st = stats.get(s["type"])
        if st:
            print(
                f"  {s['type']:20s} {st['rows']:5d} rows, "
                f"{st['pmg_id']:5d} with pmg_id, "
                f"{st['member_id']:5d} with member_id"
            )
    total_pmg = sum(1 for r in out_rows if r["pmg_id"])
    total_mid = sum(1 for r in out_rows if r["member_id"])
    print(
        f"  {'TOTAL':20s} {len(out_rows):5d} rows, "
        f"{total_pmg:5d} with pmg_id, "
        f"{total_mid:5d} with member_id"
    )
    print(f"\n  Match breakdown:")
    for status_name in [
        "pmg_code", "pmg_code_name", "member_name", "manual",
        "ambiguous", "no_match", "empty_member",
    ]:
        cnt = match_status_counts.get(status_name, 0)
        if cnt:
            print(f"    {status_name:16s}: {cnt}")
    if manual_choices:
        print(f"\n  Manual choices made ({len(manual_choices)} unique names):")
        for name, mid in sorted(manual_choices.items()):
            if mid:
                print(f"    '{name}' -> {mid}")


if __name__ == "__main__":
    main()
